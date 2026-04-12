#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import scipy.io as sio
from openpyxl import load_workbook


DEFAULT_RECON_ROOT = Path("/Volumes/FS_PROJETS/MatTechLab/yiwei.jia/recon_results")
DEFAULT_SCAN_LOG = Path("/Users/cag/Library/CloudStorage/OneDrive-HESSO/09_scan_log/MREye2-0.xlsx")
DEFAULT_SHEET_NAME = "Acquisition"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "emask_qc_reports"
EXPECTED_MASK_SIZES = {
    "T1w": 371817,
    "T2w": 646618,
}


@dataclass
class QcRow:
    subject: str
    sequence: str
    mid: str
    expected_size: int
    actual_size: int | None
    status: str
    mask_file: str
    note: str


def parse_subjects(raw_subjects: str | None) -> set[str] | None:
    if not raw_subjects:
        return None
    subjects = set()
    for token in raw_subjects.split(","):
        token = token.strip()
        if not token:
            continue
        if token.isdigit():
            subjects.add(f"{int(token):04d}")
        else:
            raise ValueError(f"Invalid subject token: {token}")
    return subjects or None


def parse_scan_log(scan_log: Path, sheet_name: str, selected_subjects: set[str] | None) -> dict[str, dict[str, str]]:
    workbook = load_workbook(scan_log, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    subject_map: dict[str, dict[str, str]] = {}

    for row in worksheet.iter_rows(min_row=3, values_only=True):
        if not row or row[0] is None:
            continue
        subject = f"{int(row[0]):04d}"
        if selected_subjects is not None and subject not in selected_subjects:
            continue

        mids = re.findall(r"(MID\d+)-(Track|T1w|T2w)", str(row[1] or ""))
        if not mids:
            continue

        mapped = {label: mid for mid, label in mids}
        if "T1w" in mapped or "T2w" in mapped:
            subject_map[subject] = mapped

    return subject_map


def load_mask_size(mask_file: Path) -> int:
    mat = sio.loadmat(mask_file)
    if "array" not in mat:
        raise RuntimeError(f"Missing 'array' in MAT file: {mask_file}")
    return int(mat["array"].size)


def qc_one_mask(subject: str, sequence: str, mid: str, criterion: str, recon_root: Path, tolerance: int) -> QcRow:
    expected_size = EXPECTED_MASK_SIZES[sequence]
    mask_file = (
        recon_root
        / subject
        / f"{mid}_recon"
        / "T1_LIBRE_Binning"
        / "et_masks"
        / f"subject_{mid}_mask_clean_{criterion}.mat"
    )

    if not mask_file.exists():
        return QcRow(
            subject=subject,
            sequence=sequence,
            mid=mid,
            expected_size=expected_size,
            actual_size=None,
            status="MISSING",
            mask_file=str(mask_file),
            note="mask file not found",
        )

    actual_size = load_mask_size(mask_file)
    if abs(actual_size - expected_size) <= tolerance:
        status = "OK"
        note = ""
    else:
        status = "SIZE_MISMATCH"
        note = f"expected around {expected_size}, got {actual_size}"

    return QcRow(
        subject=subject,
        sequence=sequence,
        mid=mid,
        expected_size=expected_size,
        actual_size=actual_size,
        status=status,
        mask_file=str(mask_file),
        note=note,
    )


def write_reports(rows: list[QcRow], output_dir: Path, criterion: str) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_criterion = criterion.replace(".", "p")
    csv_path = output_dir / f"s3_emask_qc_crit{safe_criterion}_{stamp}.csv"
    txt_path = output_dir / f"s3_emask_qc_crit{safe_criterion}_{stamp}.txt"

    with csv_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["subject", "sequence", "mid", "expected_size", "actual_size", "status", "mask_file", "note"])
        for row in rows:
            writer.writerow(
                [
                    row.subject,
                    row.sequence,
                    row.mid,
                    row.expected_size,
                    row.actual_size if row.actual_size is not None else "",
                    row.status,
                    row.mask_file,
                    row.note,
                ]
            )

    lines = [
        f"criterion: {criterion}",
        f"checked_rows: {len(rows)}",
        f"ok_rows: {sum(row.status == 'OK' for row in rows)}",
        f"problem_rows: {sum(row.status != 'OK' for row in rows)}",
        "",
    ]
    for row in rows:
        lines.append(
            " | ".join(
                [
                    row.subject,
                    row.sequence,
                    row.mid,
                    f"expected={row.expected_size}",
                    f"actual={row.actual_size if row.actual_size is not None else 'missing'}",
                    row.status,
                    row.note,
                ]
            ).rstrip()
        )
    txt_path.write_text("\n".join(lines) + "\n")
    return csv_path, txt_path


def print_summary(rows: list[QcRow]) -> None:
    print("subject | sequence | mid | expected | actual | status")
    for row in rows:
        actual = row.actual_size if row.actual_size is not None else "missing"
        print(f"{row.subject} | {row.sequence} | {row.mid} | {row.expected_size} | {actual} | {row.status}")

    mismatches = [row for row in rows if row.status != "OK"]
    if not mismatches:
        print("\nAll checked masks are within tolerance.")
        return

    print("\nProblems:")
    for row in mismatches:
        print(f"- {row.subject} {row.sequence} {row.mid}: {row.note}")


def main() -> int:
    parser = argparse.ArgumentParser(description="S3 eMask quality check against expected T1w/T2w mask lengths.")
    parser.add_argument("--subject", help="One subject or comma-separated subjects, e.g. 0001 or 0001,0002")
    parser.add_argument("--criterion", default="0.1", help="Mask criterion suffix to check, default: 0.1")
    parser.add_argument("--tolerance", type=int, default=5, help="Allowed absolute size difference, default: 5")
    parser.add_argument("--recon-root", type=Path, default=DEFAULT_RECON_ROOT)
    parser.add_argument("--scan-log", type=Path, default=DEFAULT_SCAN_LOG)
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()

    selected_subjects = parse_subjects(args.subject)
    if not args.scan_log.exists():
        raise FileNotFoundError(f"Scan log not found: {args.scan_log}")
    if not args.recon_root.exists():
        raise FileNotFoundError(f"Recon root not found: {args.recon_root}")

    subject_map = parse_scan_log(args.scan_log, args.sheet_name, selected_subjects)
    if not subject_map:
        raise RuntimeError("No matching subjects found in scan log.")

    rows: list[QcRow] = []
    for subject in sorted(subject_map):
        mapping = subject_map[subject]
        for sequence in ("T1w", "T2w"):
            mid = mapping.get(sequence)
            if not mid:
                continue
            rows.append(qc_one_mask(subject, sequence, mid, args.criterion, args.recon_root, args.tolerance))

    print_summary(rows)
    csv_path, txt_path = write_reports(rows, args.output_dir, args.criterion)
    print(f"\nCSV report: {csv_path}")
    print(f"TXT report: {txt_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
